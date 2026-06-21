"""NewRET core: CDD '4G CDD' sheet -> RETConfigWDTInternal (Internal sheet).

Shared by generate_ret.py (CLI), generate_ret_text.py (CLI) and ret_gui.py.

Simpler than ../01.RET: no 3G Installation Design sheet, no Local Cell ID band
lookup, and a FIXED set of 4 RET devices per sector (mapping.json -> devices).
Band (X) and sector (Y) come straight from CellName (New)[Key]:
    X = LEFT(RIGHT(CellName, 2), 1)   (band letter)
    Y = RIGHT(CellName, 1)            (sector letter)
All transformation rules live in mapping.json; styling comes from the template.
"""
import json
import re
from collections import defaultdict
from copy import copy

from openpyxl import load_workbook

# Output column headers (target order), same 8 columns as the WDT Internal sheet.
HEADERS = [
    "Site Name",
    "RRU Name",
    "RRU CN",
    "RRU SRN",
    "RRU SN",
    "RCU Coloring",
    "RCU Tilt",
    "Device Name",
]


def load_mapping(path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def _norm_header(s):
    """Normalize a header cell for tolerant matching (case/whitespace-insensitive)."""
    if s is None:
        return ""
    return " ".join(str(s).split()).strip().lower()


def _find_index(norm_headers, targets):
    """Index of the first header matching any target (exact, then startswith)."""
    targets = [t for t in targets if t]
    for t in targets:
        if t in norm_headers:
            return norm_headers.index(t)
    for i, h in enumerate(norm_headers):
        if h and any(h.startswith(t) for t in targets):
            return i
    return None


def _locate_header(all_rows, key_targets, max_scan=10):
    """Find the header row by locating the key column; return (row_index, norm_headers).

    The 4G CDD sheet has a category row above the real header row, so the header
    is not row 0. Scans the first ``max_scan`` rows; falls back to row 0.
    """
    for i, row in enumerate(all_rows[:max_scan]):
        norm = [_norm_header(h) for h in row]
        if _find_index(norm, key_targets) is not None:
            return i, norm
    return 0, [_norm_header(h) for h in (all_rows[0] if all_rows else ())]


def _to_float(v):
    """Best-effort float; None for blank/'-' or unparseable."""
    if v in (None, "", "-"):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def resolve_sector(Y, mapping, offset=0):
    """Map the sector letter/digit Y (+ offset) to (sector_id, rru_srn), or None.

    Per mapping.json -> sector_rule: a digit 1-9 is that base sector number; a
    single letter A-I maps to 1-9 (A=S1 ... I=S9). ``offset`` (the co-located
    offset, 0 for the NE's own site) is added to that base number. sector_id =
    'S{n}', rru_srn = srn_base + (n - 1). Returns None for anything outside
    1..max_sectors.
    """
    rule = mapping.get("sector_rule", {})
    base = rule.get("srn_base", 60)
    max_sectors = rule.get("max_sectors", 9)
    Y = str(Y).strip()
    n = None
    if Y.isdigit():
        n = int(Y)
    elif len(Y) == 1 and Y.upper().isalpha():
        n = ord(Y.upper()) - ord("A") + 1
    if n is None:
        return None
    n += offset
    if not (1 <= n <= max_sectors):
        return None
    return f"S{n}", base + (n - 1)


def list_sheets(cdd_path):
    """Return sheet names in the CDD workbook."""
    wb = load_workbook(cdd_path, read_only=True)
    try:
        return wb.sheetnames
    finally:
        wb.close()


def _resolve_columns(all_rows, mapping):
    """Locate CDD columns by header name. Returns (header_row_index, col_map)."""
    hdr_cfg = mapping.get("source", {}).get("headers", {})
    wanted = {
        "site_new": hdr_cfg.get("site_new", "SiteName (RRU Location)_New"),
        "ne_name": hdr_cfg.get("ne_name", "NEName_New"),
        "ne_id": hdr_cfg.get("ne_id", "Ne ID (New)"),
        "cell_name": hdr_cfg.get("cell_name", "CellName (New)[Key]"),
        "e_tilt": hdr_cfg.get("e_tilt", "E_TILT"),
        "bbu_cluster": hdr_cfg.get("bbu_cluster", "BBU Cluster"),
    }
    hdr_idx, norm = _locate_header(all_rows, [_norm_header(wanted["cell_name"])])
    col = {f: _find_index(norm, [_norm_header(name)]) for f, name in wanted.items()}
    return hdr_idx, col, wanted


def list_bbu_clusters(cdd_path, sheet, mapping):
    """Return the sorted distinct BBU Cluster values present in the sheet."""
    wb = load_workbook(cdd_path, data_only=True, read_only=True)
    try:
        ws = wb[sheet]
        all_rows = list(ws.iter_rows(min_row=1, values_only=True))
    finally:
        wb.close()
    if not all_rows:
        return []
    hdr_idx, col, _ = _resolve_columns(all_rows, mapping)
    j = col.get("bbu_cluster")
    cj = col.get("cell_name")
    if j is None:
        return []
    found = set()
    for r in all_rows[hdr_idx + 1:]:
        if not r or (cj is not None and (len(r) <= cj or r[cj] is None)):
            continue
        v = r[j] if len(r) > j else None
        if v not in (None, ""):
            found.add(str(v).strip())
    return sorted(found)


def build_rows(cdd_path, sheet, mapping, clusters=None):
    """Parse the 4G CDD sheet and produce output rows.

    ``clusters`` (optional): an iterable of BBU Cluster values; when given, only
    sectors whose BBU Cluster is in the set are emitted. None/empty = all.

    Returns (rows, skipped, sector_count, site_index) where:
      rows         = list of 8-value lists (matching HEADERS)
      skipped      = list of (cell_name, reason)
      sector_count = number of distinct (site, sector) groups emitted
      site_index   = {site_new: {"prefix": ..., "tilts": {(sector_id, pos): tilt}}}
                     for the MML text feature.
    """
    devices = mapping["devices"]
    rru_cn = mapping["constants"]["rru_cn"]
    rru_sn = mapping["constants"]["rru_sn"]
    field_rules = mapping.get("field_rules", {})
    # Site Name(*) comes from this CDD field (default NEName_New).
    site_name_source = field_rules.get("site_name", {}).get("source", "ne_name")
    # RRU Name(*) prefix = the part of this field before the first of these delimiters.
    rru_prefix_source = field_rules.get("rru_name", {}).get("prefix_source", "site_new")
    rru_prefix_delims = field_rules.get("rru_name", {}).get("prefix_delimiters", ["_"])
    sector_rule = mapping.get("sector_rule", {})
    match_len = sector_rule.get("match_prefix_len", 8)
    colocated_offset = sector_rule.get("colocated_offset", 3)
    cluster_filter = {str(c).strip() for c in clusters} if clusters else None

    wb = load_workbook(cdd_path, data_only=True, read_only=True)
    try:
        ws = wb[sheet]
        all_rows = list(ws.iter_rows(min_row=1, values_only=True))
    finally:
        wb.close()

    hdr_idx, col, wanted = _resolve_columns(all_rows, mapping)
    required = ("site_new", "cell_name")
    missing = [wanted[f] for f in required if col[f] is None]
    if missing:
        found = [str(h) for h in (all_rows[hdr_idx] if all_rows else ()) if h is not None]
        raise ValueError(
            "Sheet '%s': could not find required column(s) by header name: %s.\n"
            "Headers found in the sheet:\n  %s\n"
            "Check mapping.json -> source.headers."
            % (sheet, ", ".join('"%s"' % m for m in missing), "\n  ".join(found))
        )

    def _get(r, field):
        j = col[field]
        return r[j] if (j is not None and len(r) > j) else None

    by_sector = defaultdict(dict)   # (site_new, Y) -> {X: e_tilt}
    sector_ne_id = {}               # (site_new, Y) -> Ne ID (first non-blank)
    sector_ne_name = {}             # (site_new, Y) -> NEName_New (first non-blank)
    sector_offset = {}              # (site_new, Y) -> 0 (own site) or colocated_offset
    sector_order = []
    seen = set()
    skipped = []

    for r in all_rows[hdr_idx + 1:]:
        if not r or _get(r, "cell_name") is None:
            continue
        cell = str(_get(r, "cell_name")).strip()
        if len(cell) < 2:
            skipped.append((cell, "too short"))
            continue
        site_new = _get(r, "site_new")
        site_new = str(site_new).strip() if site_new is not None else ""
        if not site_new:
            skipped.append((cell, "blank SiteName_New"))
            continue
        # BBU Cluster filter (when a selection is active).
        if cluster_filter is not None:
            bbu = _get(r, "bbu_cluster")
            bbu = str(bbu).strip() if bbu is not None else ""
            if bbu not in cluster_filter:
                continue
        X = cell[-2]   # band letter
        Y = cell[-1]   # sector letter/digit
        # Own site (cell belongs to this NE) keeps its base sector number; a
        # co-located neighbour (LEFT(NEName,N) != LEFT(CellName,N)) is shifted by
        # colocated_offset (S1->S4, S2->S5, S3->S6).
        ne_name_val = _get(r, "ne_name")
        ne_name_val = str(ne_name_val).strip() if ne_name_val is not None else ""
        is_own = ne_name_val[:match_len].upper() == cell[:match_len].upper()
        offset = 0 if is_own else colocated_offset
        if resolve_sector(Y, mapping, offset) is None:
            skipped.append((cell, f"unknown sector Y={Y}"))
            continue
        key = (site_new, Y)
        if key not in seen:
            seen.add(key)
            sector_order.append(key)
            sector_offset[key] = offset
        ne_id = _get(r, "ne_id")
        if ne_id not in (None, "") and key not in sector_ne_id:
            sector_ne_id[key] = ne_id
        nn = _get(r, site_name_source)
        if nn not in (None, "") and key not in sector_ne_name:
            sector_ne_name[key] = str(nn).strip()
        # E_TILT per band; first non-blank wins so a blank row can't clobber it.
        t = _to_float(_get(r, "e_tilt"))
        if X not in by_sector[key] or (by_sector[key].get(X) is None and t is not None):
            by_sector[key][X] = t

    rows = []
    site_index = {}
    for key in sector_order:
        site_new, Y = key
        sector_id, rru_srn = resolve_sector(Y, mapping, sector_offset.get(key, 0))
        present = by_sector[key]            # {X: e_tilt} for this sector
        ne_id = sector_ne_id.get(key, "")
        # Site Name(*) = NEName_New (fallback to SiteName_New if blank).
        site_name = sector_ne_name.get(key, "") or site_new
        # RRU Name(*) prefix = part of the configured source field before the first
        # delimiter char (default SiteName (RRU Location)_New).
        rru_prefix = {"site_new": site_new, "ne_name": site_name,
                      "site_name": site_name}.get(rru_prefix_source, site_new)
        for d in rru_prefix_delims:
            rru_prefix = rru_prefix.split(d)[0]
        site_entry = site_index.setdefault(
            site_new, {"prefix": f"{site_new}_{ne_id}".rstrip("_"), "tilts": {}}
        )

        for pos, dev in enumerate(devices):
            t = present.get(dev["tilt_band"])
            if t is None:
                t = present.get(dev.get("tilt_fallback"))
            rcu_tilt = round((t or 0.0) * 10)
            rru_name = f"{rru_prefix}_{dev['band_token']}_{sector_id}{dev['slot_suffix']}"
            site_entry["tilts"][(sector_id, pos)] = rcu_tilt
            rows.append([
                site_name,
                rru_name,
                rru_cn,
                rru_srn,
                rru_sn,
                dev["color"],
                rcu_tilt,
                rru_name,   # column H = column B (Device Name = RRU Name)
            ])

    return rows, skipped, len(sector_order), site_index


def _find_template_header_row(ws, headers, max_scan=20):
    """Find the column-header row in the template by matching known header tokens.

    Templates may have preamble rows above the header (e.g. a Declaration note),
    so the header is not always row 1. Returns the 1-based row index; falls back
    to row 1.
    """
    targets = [_norm_header(h) for h in headers]
    max_col = ws.max_column or 1
    best_row, best_score = 1, 0
    for i in range(1, min(ws.max_row or 1, max_scan) + 1):
        cells = [_norm_header(ws.cell(row=i, column=c).value) for c in range(1, max_col + 1)]
        score = sum(1 for t in targets if t and any(t in cell for cell in cells))
        if score > best_score:
            best_score, best_row = score, i
    return best_row if best_score >= 2 else 1


def write_output(template_path, target_sheet, rows, output_path):
    """Write rows into a copy of the template, preserving its header rows and the
    styling of its first data row.

    The header row is detected (it may sit below preamble rows), so everything up
    to and including the header is kept and data is written from the next row.
    """
    out_wb = load_workbook(template_path)
    out_ws = out_wb[target_sheet]

    header_row = _find_template_header_row(out_ws, HEADERS)
    data_start = header_row + 1

    style_row = data_start if (out_ws.max_row or 0) >= data_start else header_row
    template_row_styles = []
    for c in range(1, out_ws.max_column + 1):
        cell = out_ws.cell(row=style_row, column=c)
        template_row_styles.append({
            "font": copy(cell.font),
            "fill": copy(cell.fill),
            "border": copy(cell.border),
            "alignment": copy(cell.alignment),
            "number_format": cell.number_format,
        })

    # Clear existing data rows only (keep preamble + header).
    if (out_ws.max_row or 0) >= data_start:
        out_ws.delete_rows(data_start, out_ws.max_row - data_start + 1)

    for i, values in enumerate(rows):
        out_row = data_start + i
        for col_idx, v in enumerate(values, start=1):
            c = out_ws.cell(row=out_row, column=col_idx, value=v)
            if col_idx - 1 < len(template_row_styles):
                s = template_row_styles[col_idx - 1]
                c.font = copy(s["font"])
                c.fill = copy(s["fill"])
                c.border = copy(s["border"])
                c.alignment = copy(s["alignment"])
                c.number_format = s["number_format"]

    out_wb.save(output_path)


# --------------------------------------------------------------------------
# RET_template.txt -> RET_output.txt (MML script) conversion.
#
# Rewrites three things in an "ADD RET / MOD RETTILT" MML template using the
# RET_input.txt serials and the CDD-derived rows (build_rows):
#   * DEVICENAME : replace the site-prefix tokens with {SiteName_New}_{Ne ID}.
#   * SERIALNO   : take the input serial matched by (CTRLSRN, RIGHT(serial, N)).
#   * TILT       : RCU Tilt of the same-DEVICENO ADD RET device, matched
#                  positionally (CTRLSRN -> sector, device order within sector).
# All rules/field names live in mapping.json -> text_config.
# --------------------------------------------------------------------------

_RE_DEVICENO = re.compile(r"DEVICENO=\s*(\d+)")
_RE_DEVICENAME = re.compile(r'DEVICENAME="([^"]*)"')
_RE_CTRLSRN = re.compile(r"CTRLSRN=\s*(\d+)")
_RE_SERIALNO = re.compile(r'SERIALNO="([^"]*)"')
_RE_TILT = re.compile(r"TILT=\s*(-?\d+)")


def parse_ret_input(input_path, mapping):
    """Parse RET_input.txt -> (site, serials)."""
    with open(input_path, encoding="utf-8") as f:
        text = f.read()
    return parse_ret_input_text(text, mapping, source=input_path)


def parse_ret_input_text(text, mapping, source="pasted input"):
    """Parse in-memory RET input text -> (site, serials).

    Line 1 (first non-blank) is the site name. Each remaining line is split on
    whitespace; ``serials`` maps (CTRLSRN, RIGHT(serial, suffix_len)) -> full
    serial, the key used to rewrite each template SERIALNO.
    """
    cfg = mapping.get("text_config", {}).get("input", {})
    srn_col = cfg.get("ctrlsrn_column", 1)
    ser_col = cfg.get("serial_column", 7)
    suf_len = cfg.get("serial_suffix_len", 3)

    site = None
    serials = {}
    for ln in text.splitlines():
        if not ln.strip():
            continue
        if site is None:
            site = ln.strip()
            continue
        parts = ln.split()
        if len(parts) <= max(srn_col, ser_col):
            continue
        srn = parts[srn_col].strip()
        serial = parts[ser_col].strip()
        serials[(srn, serial[-suf_len:])] = serial
    if site is None:
        raise ValueError("RET input is empty: %s" % source)
    return site, serials


def _site_tilt_index(site_index, site):
    """Return (new_prefix, tilt_by_sector_pos) for ``site`` from build_rows."""
    entry = site_index.get(site)
    if entry is None:
        available = ", ".join(sorted(site_index))
        raise ValueError(
            "Site %r (from RET input) was not found in the CDD output.\n"
            "Sites available in the CDD: %s" % (site, available)
        )
    return entry["prefix"], entry["tilts"]


def build_text_output(template_path, input_path, cdd_path, sheet, mapping,
                      input_text=None, clusters=None):
    """Produce the rewritten MML text. Returns (text, warnings)."""
    if input_text is not None and input_text.strip():
        site, serials = parse_ret_input_text(input_text, mapping)
    else:
        site, serials = parse_ret_input(input_path, mapping)
    _rows, _, _, site_index = build_rows(cdd_path, sheet, mapping, clusters=clusters)
    new_prefix, tilt_by_sector_pos = _site_tilt_index(site_index, site)

    # Map CTRLSRN -> sector_id by inverting the sector_rule over 1..max_sectors.
    max_sectors = mapping.get("sector_rule", {}).get("max_sectors", 9)
    srn_to_sector = {}
    for n in range(1, max_sectors + 1):
        sid, srn = resolve_sector(str(n), mapping)
        srn_to_sector[str(srn)] = sid

    tcfg = mapping.get("text_config", {}).get("template", {})
    prefix_tokens = tcfg.get("prefix_token_count", 2)
    add_prefix = tcfg.get("add_line_prefix", "ADD RET")
    tilt_prefix = tcfg.get("tilt_line_prefix", "MOD RETTILT")
    suf_len = mapping.get("text_config", {}).get("input", {}).get("serial_suffix_len", 3)

    with open(template_path, encoding="utf-8") as f:
        tmpl_lines = f.readlines()

    warnings = []

    # Pass 1: DEVICENO -> RCU Tilt, using each ADD RET line's CTRLSRN (sector)
    # and its order within that sector (positional match).
    deviceno_tilt = {}
    add_pos = defaultdict(int)
    for ln in tmpl_lines:
        if not ln.lstrip().startswith(add_prefix):
            continue
        m_no, m_srn = _RE_DEVICENO.search(ln), _RE_CTRLSRN.search(ln)
        if not (m_no and m_srn):
            continue
        srn = m_srn.group(1)
        pos = add_pos[srn]
        add_pos[srn] += 1
        sector_id = srn_to_sector.get(srn)
        tilt = tilt_by_sector_pos.get((sector_id, pos))
        if tilt is None:
            warnings.append(
                "No RCU Tilt for DEVICENO=%s (CTRLSRN=%s, sector=%s, pos=%d)"
                % (m_no.group(1), srn, sector_id, pos)
            )
        deviceno_tilt[m_no.group(1)] = tilt

    # Pass 2: rewrite lines in place, preserving everything else verbatim.
    out_lines = []
    for ln in tmpl_lines:
        stripped = ln.lstrip()
        if stripped.startswith(add_prefix):
            def _sub_devicename(m):
                tokens = m.group(1).split("_")
                suffix = "_".join(tokens[prefix_tokens:])
                return 'DEVICENAME="%s"' % (new_prefix + "_" + suffix)

            ln = _RE_DEVICENAME.sub(_sub_devicename, ln)

            m_srn = _RE_CTRLSRN.search(ln)
            if m_srn:
                srn = m_srn.group(1)

                def _sub_serial(m):
                    suffix = m.group(1)[-suf_len:]
                    new = serials.get((srn, suffix))
                    if new is None:
                        warnings.append(
                            "No input serial for CTRLSRN=%s suffix=%s" % (srn, suffix)
                        )
                        return m.group(0)
                    return 'SERIALNO="%s"' % new

                ln = _RE_SERIALNO.sub(_sub_serial, ln)
            out_lines.append(ln)
        elif stripped.startswith(tilt_prefix):
            m_no = _RE_DEVICENO.search(ln)
            if m_no:
                tilt = deviceno_tilt.get(m_no.group(1))
                if tilt is not None:
                    ln = _RE_TILT.sub("TILT=%d" % int(round(float(tilt))), ln)
                else:
                    warnings.append("No tilt for MOD RETTILT DEVICENO=%s" % m_no.group(1))
            out_lines.append(ln)
        else:
            out_lines.append(ln)

    return "".join(out_lines), warnings


def write_text_output(template_path, input_path, cdd_path, sheet, mapping, output_path,
                      input_text=None, clusters=None):
    """build_text_output + write to ``output_path``. Returns warnings."""
    text, warnings = build_text_output(
        template_path, input_path, cdd_path, sheet, mapping,
        input_text=input_text, clusters=clusters,
    )
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(text)
    return warnings
